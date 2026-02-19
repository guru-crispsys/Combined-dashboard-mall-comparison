import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from urllib.parse import urlparse
import re


def create_mall_excel_export(
    scraped_df=None,
    structured_data=None,
    llm_json=None,
    input_url="",
    output_buffer=None
):
    """
    Create an Excel file with 7 tabs:
    1. Mall Meta Data
    2. Existing Tennent Research
    3. Coming Soon Tennent Research
    4. Vacated Shops
    5. AI Analysis Report
    6. Facebook Scratch
    7. Instagram Scratch
    
    Args:
        scraped_df: DataFrame with scraped tenant data
        structured_data: Comparison data structure
        llm_json: LLM analysis results
        input_url: URL(s) used for scraping
        output_buffer: BytesIO buffer to write to (if None, creates new)
    """
    from io import BytesIO
    
    if output_buffer is None:
        output_buffer = BytesIO()
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Extract metadata - first try from llm_json, then from input_url
    metadata = _extract_metadata(input_url, llm_json)

    # Add scrape statistics into metadata for display in Mall Meta Data tab
    if scraped_df is not None and not scraped_df.empty and 'source' in scraped_df.columns:
        source_series = scraped_df['source'].astype(str).str.lower()
        website_count = int(source_series.str.contains('website', na=False).sum())
        facebook_count = int(source_series.str.contains('facebook', na=False).sum())
        instagram_count = int(source_series.str.contains('instagram', na=False).sum())
        total_scraped = int(len(scraped_df))

        metadata["scraped_website_count"] = website_count
        metadata["scraped_facebook_count"] = facebook_count
        metadata["scraped_instagram_count"] = instagram_count
        metadata["scraped_total_count"] = total_scraped
    
    # Extract coming soon shops from website text using AI
    coming_soon_shops = []
    if input_url:
        try:
            # Try to read extracted text files if available
            import os
            extracted_text = ""
            website_url = ""
            
            # Extract website URL from input_url (not Facebook/Instagram)
            if input_url:
                import re
                from urllib.parse import urlparse
                url_pattern = re.compile(r"https?://[^\s,\n]+")
                urls = url_pattern.findall(input_url)
                for url in urls:
                    url_lower = url.lower()
                    if 'facebook.com' not in url_lower and 'instagram.com' not in url_lower and 'fb.com' not in url_lower and 'instagr.am' not in url_lower:
                        website_url = url
                        break
            
            # Try to read from last_extracted_text_path.txt
            if os.path.exists("last_extracted_text_path.txt"):
                with open("last_extracted_text_path.txt", "r", encoding="utf-8") as f:
                    text_file_path = f.read().strip()
                    if os.path.exists(text_file_path):
                        with open(text_file_path, "r", encoding="utf-8") as text_file:
                            # Skip header lines and get the actual text content
                            lines = text_file.readlines()
                            # Find the separator line (=====)
                            start_idx = 0
                            for i, line in enumerate(lines):
                                if "=" * 80 in line or "=" * 40 in line:
                                    start_idx = i + 1
                                    break
                            extracted_text = "\n".join(lines[start_idx:])
            
            # If we have extracted text, use AI to extract coming soon shops
            if extracted_text and len(extracted_text.strip()) > 50:
                from llm_engine import extract_coming_soon_shops_from_text
                print("Extracting coming soon shops from website text using AI...")
                coming_soon_shops = extract_coming_soon_shops_from_text(extracted_text, url=website_url)
                print(f"Found {len(coming_soon_shops)} coming soon shops")
        except Exception as e:
            print(f"Warning: Failed to extract coming soon shops: {e}")
            import traceback
            traceback.print_exc()
    
    # Fetch SERP news/blogs for mall (mall name + address from main UI)
    # Use same query style as terminal so Excel "Google SERP Scratch" matches: title-case mall name, full results.
    google_search_results = []
    try:
        mall_name = metadata.get("mall_name") or ""
        address = metadata.get("address") or ""
        if (not mall_name or mall_name == "Not Available") or (not address or address == "Not Available"):
            # Try shared_dashboard_input.json from main UI
            _root = __file__
            for _ in range(2):
                _root = __import__("os").path.dirname(_root)
            _shared = __import__("pathlib").Path(_root) / "shared_dashboard_input.json"
            if _shared.exists():
                _data = __import__("json").loads(_shared.read_text(encoding="utf-8"))
                if not mall_name or mall_name == "Not Available":
                    mall_name = (_data.get("mall_name") or "").strip()
                if not address or address == "Not Available":
                    address = (_data.get("address") or "").strip()
        mall_name = mall_name if mall_name and mall_name != "Not Available" else ""
        address = address if address and address != "Not Available" else ""
        # Normalize for SERP: title-case mall name so we get same news-style results as terminal (e.g. "Plaza Frontenac")
        if mall_name:
            mall_name = " ".join(w.capitalize() for w in mall_name.split())
        if mall_name or address:
            from serp_news_scraper import fetch_mall_news
            google_search_results = fetch_mall_news(mall_name, address, max_results=15)
    except Exception as e:
        pass

    # Create tabs
    _create_meta_data_tab(wb, metadata)
    _create_existing_tenants_tab(wb, scraped_df, structured_data, google_search_results=google_search_results)
    _create_coming_soon_tab(wb, structured_data, coming_soon_shops=coming_soon_shops)
    _create_vacated_shops_tab(wb, structured_data)
    _create_ai_analysis_tab(wb, llm_json, structured_data)
    _create_facebook_scratch_tab(wb, scraped_df)
    _create_instagram_scratch_tab(wb, scraped_df)
    _create_serp_scratch_tab(wb, google_search_results)
    
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def create_existing_tenant_research_only_export(
    scraped_df=None,
    structured_data=None,
    input_url="",
    output_buffer=None,
):
    """
    Create an Excel file with only the Existing Tenant Research tab (website + FB + IG data).
    """
    from io import BytesIO

    if output_buffer is None:
        output_buffer = BytesIO()

    wb = Workbook()
    wb.remove(wb.active)
    # For existing-tenant-only export, fetch SERP so column L/M can be filled
    google_search_results = []
    try:
        metadata = _extract_metadata(input_url, None)
        mall_name = (metadata.get("mall_name") or "").strip()
        address = (metadata.get("address") or "").strip()
        if (not mall_name or mall_name == "Not Available") or (not address or address == "Not Available"):
            _root = __file__
            for _ in range(2):
                _root = __import__("os").path.dirname(_root)
            _shared = __import__("pathlib").Path(_root) / "shared_dashboard_input.json"
            if _shared.exists():
                _data = __import__("json").loads(_shared.read_text(encoding="utf-8"))
                if not mall_name or mall_name == "Not Available":
                    mall_name = (_data.get("mall_name") or "").strip()
                if not address or address == "Not Available":
                    address = (_data.get("address") or "").strip()
        mall_name = mall_name if mall_name and mall_name != "Not Available" else ""
        address = address if address and address != "Not Available" else ""
        if mall_name:
            mall_name = " ".join(w.capitalize() for w in mall_name.split())
        if mall_name or address:
            from serp_news_scraper import fetch_mall_news
            google_search_results = fetch_mall_news(mall_name, address, max_results=15)
    except Exception:
        pass
    _create_existing_tenants_tab(wb, scraped_df, structured_data, google_search_results=google_search_results)
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def _score_post_for_tenant(post_text: str, tenant_name: str) -> int:
    """
    Score how strongly a Facebook/Instagram post matches a tenant name.
    Higher score = stronger, more specific match. 0 means "no match".
    This lets us assign each post to the *best* tenant when multiple names appear.
    """
    if not post_text or not tenant_name:
        return 0

    post_lower = str(post_text).lower().strip()
    tenant_lower = str(tenant_name).lower().strip()
    tenant_compact = tenant_lower.replace(" ", "")

    if not post_lower or not tenant_lower:
        return 0

    score = 0

    # Important words from tenant name (drop generic ones)
    tenant_words = [
        w
        for w in tenant_lower.split()
        if len(w) > 2 and w not in ["the", "and", "or", "for", "with", "mall", "store", "shop"]
    ]
    if not tenant_words:
        tenant_words = [tenant_lower]

    # Whole name match (with and without spaces) – strongest signal
    if len(tenant_lower) >= 3 and re.search(r"\b" + re.escape(tenant_lower) + r"\b", post_lower):
        score += 100
    if len(tenant_compact) >= 3 and tenant_compact in post_lower.replace(" ", ""):
        score += 80

    # Word-level matches for important words
    for w in tenant_words:
        if len(w) >= 3 and re.search(r"\b" + re.escape(w) + r"\b", post_lower):
            score += 20

    # Conservative partial match (e.g., "Shake Shack" vs "Shake Shack Bellevue Square")
    if len(tenant_lower) >= 4 and tenant_lower in post_lower:
        score += 10

    return score


def _match_post_to_tenant(post_text, tenant_name):
    """
    Backwards-compatible wrapper: treat any positive score as a match.
    Prefer using _score_post_for_tenant when choosing between multiple tenants.
    """
    return _score_post_for_tenant(post_text, tenant_name) > 0


def _assign_serp_results_to_tenants(google_search_results, tenant_data):
    """
    Assign each SERP news/blog result to the best-matching tenant row (like Facebook/Instagram).
    Returns list of (google_info_text, google_url_text) per tenant, same length as tenant_data.
    """
    from serp_news_scraper import format_news_for_excel
    per_tenant = [[] for _ in tenant_data]
    for item in google_search_results:
        title = (item.get("title") or "").strip()
        snippet = (item.get("snippet") or "").strip()
        combined_text = f"{title}\n{snippet}".strip()
        if not combined_text:
            combined_text = title or snippet
        best_idx = None
        best_score = 0
        for idx, tenant in enumerate(tenant_data):
            tenant_name = tenant.get("name", "") or ""
            if not tenant_name:
                continue
                score = _score_post_for_tenant(combined_text, tenant_name)
                if score > best_score:
                    best_score = score
                    best_idx = idx
        # Only assign when there is a positive match; otherwise keep item only in SERP Scratch
        if best_idx is not None and best_score > 0:
            per_tenant[best_idx].append(item)
    # Format text and URLs for each tenant
    out = []
    for items in per_tenant:
        text, urls = format_news_for_excel(items)
        out.append((text, urls))
    return out


def _process_serp_for_tenants(google_search_results, tenant_data):
    """
    Process SERP results for Existing Tenant Research: use AI to extract/clean info and match
    to tenants when possible; otherwise fall back to score-based assignment.
    Returns list of (google_info_text, google_url_text) per tenant, same length as tenant_data.
    """
    from serp_news_scraper import format_news_for_excel
    if not google_search_results or not tenant_data:
        return [( "", "" )] * len(tenant_data) if tenant_data else []

    tenant_names = [str(t.get("name") or "").strip() for t in tenant_data if t.get("name")]
    try:
        from llm_engine import extract_serp_with_ai
        extracted = extract_serp_with_ai(google_search_results, tenant_names)
    except Exception:
        extracted = []

    if not extracted:
        return _assign_serp_results_to_tenants(google_search_results, tenant_data)

    # Assign by matched_tenant: find tenant index by name (exact or first containing match).
    # If no tenant is matched, do NOT assign; such rows will still appear in the
    # Google SERP Scratch tab but not in Existing Tenant Research.
    per_tenant = [[] for _ in tenant_data]
    for item in extracted:
        matched = item.get("matched_tenant")
        if matched and matched.strip():
            matched_clean = matched.strip()
            target_idx = None
            for idx, t in enumerate(tenant_data):
                name = (t.get("name") or "").strip()
                if name and (name == matched_clean or matched_clean in name or name in matched_clean):
                    target_idx = idx
                    break
            if target_idx is not None:
                per_tenant[target_idx].append(item)

    out = []
    for items in per_tenant:
        text, urls = format_news_for_excel(items)
        out.append((text, urls))
    return out


def _extract_metadata(input_url, llm_json=None):
    """Extract mall metadata from URL(s) and llm_json"""
    # Start with defaults
    metadata = {
        "mall_name": "Not Available",
        "address": "Not Available",
        "official_website": "Not Available",
        "facebook_link": "Not Available",
        "instagram_link": "Not Available",
        "hashtags": "#Shopping #CityLife",
        "research_run_date": datetime.now().strftime("%Y-%m-%d")
    }
    
    # First, try to get metadata from llm_json if available
    if llm_json and isinstance(llm_json, dict):
        llm_metadata = llm_json.get("metadata", {})
        if llm_metadata:
            # Extract official_website and clean it (remove Facebook/Instagram URLs if present)
            official_website = llm_metadata.get("official_website", metadata["official_website"])
            if official_website and official_website != "Not Available":
                # Check if it contains multiple URLs (Facebook/Instagram)
                # Note: re is already imported at the top of the file
                url_pattern = re.compile(r"https?://[^\s,\n]+")
                urls_in_field = url_pattern.findall(str(official_website))
                # Extract only website URL (not Facebook/Instagram)
                website_only = [u for u in urls_in_field if 'facebook.com' not in u.lower() and 'instagram.com' not in u.lower() and 'fb.com' not in u.lower() and 'instagr.am' not in u.lower()]
                if website_only:
                    official_website = website_only[0]
                elif urls_in_field:
                    # If all URLs are Facebook/Instagram, set to Not Available
                    official_website = "Not Available"
            
            metadata.update({
                "mall_name": llm_metadata.get("mall_name", metadata["mall_name"]),
                "address": llm_metadata.get("address", metadata["address"]),
                "official_website": official_website,
                "facebook_link": llm_metadata.get("facebook_link", metadata["facebook_link"]),
                "instagram_link": llm_metadata.get("instagram_link", metadata["instagram_link"]),
                "hashtags": llm_metadata.get("hashtags", metadata["hashtags"]),
                # Always use current date for research_run_date, never override with LLM date
                "research_run_date": datetime.now().strftime("%Y-%m-%d")
            })
    
    # Initialize URL lists
    website_urls = []
    facebook_urls = []
    instagram_urls = []
    
    # Then, extract from input_url to fill in missing values
    if input_url:
        # Parse URLs - handle multiple formats (comma, newline, space separated)
        # Note: re is already imported at the top of the file
        url_pattern = re.compile(r"https?://[^\s,\n]+")
        urls = url_pattern.findall(input_url)
        
        # Also try splitting by comma and newline as fallback
        if not urls:
            urls = input_url.replace('\n', ',').replace('\r', ',').split(',')
            urls = [u.strip() for u in urls if u.strip() and ('http://' in u or 'https://' in u)]
        
        for url in urls:
            url = url.strip()
            if not url:
                continue
            # Ensure URL starts with http:// or https://
            if not url.startswith('http://') and not url.startswith('https://'):
                continue
            if 'facebook.com' in url.lower() or 'fb.com' in url.lower():
                facebook_urls.append(url)
            elif 'instagram.com' in url.lower() or 'instagr.am' in url.lower():
                instagram_urls.append(url)
            else:
                website_urls.append(url)
    
    # Set website URL (only if not already set from llm_json, or if llm_json has all URLs concatenated)
    # Check if official_website contains multiple URLs (Facebook/Instagram links)
    current_official_website = metadata.get("official_website", "")
    if current_official_website and current_official_website != "Not Available":
        # Check if it contains Facebook or Instagram URLs (indicates concatenation issue)
        if 'facebook.com' in current_official_website.lower() or 'instagram.com' in current_official_website.lower():
            # Extract only the website URL (first URL that's not Facebook/Instagram)
            # Note: url_pattern is defined above, but if we're here, we need to recreate it
            url_pattern = re.compile(r"https?://[^\s,\n]+")
            urls_in_field = url_pattern.findall(current_official_website)
            website_only = [u for u in urls_in_field if 'facebook.com' not in u.lower() and 'instagram.com' not in u.lower() and 'fb.com' not in u.lower()]
            if website_only:
                metadata["official_website"] = website_only[0]
            elif website_urls:
                metadata["official_website"] = website_urls[0]
            else:
                metadata["official_website"] = "Not Available"
    
    # Set website URL if not already set or if we have website URLs from input_url
    if website_urls and (not metadata.get("official_website") or metadata.get("official_website") == "Not Available"):
        metadata["official_website"] = website_urls[0]
        # Try to extract mall name from domain and URL path
        try:
            parsed = urlparse(website_urls[0])
            domain = parsed.netloc.replace("www.", "")
            
            # Try extracting from domain first (e.g., vishaalmall.com -> Vishaalmall)
            if domain:
                domain_parts = domain.split(".")
                if domain_parts:
                    domain_name = domain_parts[0].replace("-", " ").title()
                    # If domain name looks like a mall name (contains 'mall' or is meaningful)
                    if "mall" in domain_name.lower() or len(domain_name) > 3:
                        metadata["mall_name"] = domain_name
            
            # Also try extracting from URL path (e.g., /vishaal-mall/ -> Vishaal Mall)
            if (not metadata["mall_name"] or metadata["mall_name"] == "Not Available") and parsed.path:
                path_parts = [p for p in parsed.path.split("/") if p and len(p) > 2]
                if path_parts:
                    # Take the first meaningful path segment
                    path_name = path_parts[0].replace("-", " ").replace("_", " ").title()
                    if len(path_name) > 3:
                        metadata["mall_name"] = path_name
        except Exception as e:
            pass
    
    # Set Facebook URL (only if not already set from llm_json)
    if facebook_urls and (not metadata.get("facebook_link") or metadata.get("facebook_link") == "Not Available"):
        metadata["facebook_link"] = facebook_urls[0]
        # Try to extract mall name from Facebook URL
        try:
            if not metadata["mall_name"] or metadata["mall_name"] == "Not Available":
                parts = facebook_urls[0].rstrip('/').split('/')
                if parts:
                    name_part = parts[-1]
                    metadata["mall_name"] = name_part.replace("-", " ").replace(".", " ").title()
        except:
            pass
    
    # Set Instagram URL (only if not already set from llm_json)
    if instagram_urls and (not metadata.get("instagram_link") or metadata.get("instagram_link") == "Not Available"):
        metadata["instagram_link"] = instagram_urls[0]
        # Try to extract mall name from Instagram URL
        try:
            if not metadata["mall_name"] or metadata["mall_name"] == "Not Available":
                parts = instagram_urls[0].rstrip('/').split('/')
                if parts:
                    name_part = parts[-1]
                    metadata["mall_name"] = name_part.replace("-", " ").replace(".", " ").replace("_", " ").title()
        except:
            pass
    
    # Generate Instagram URL from Facebook URL if Instagram URL not available
    if metadata["instagram_link"] == "Not Available" and metadata["facebook_link"] != "Not Available":
        try:
            # Common pattern: if Facebook is /VishaalMall/, Instagram might be /vishaalmall/
            fb_name = metadata["facebook_link"].rstrip('/').split('/')[-1]
            metadata["instagram_link"] = f"https://www.instagram.com/{fb_name.lower()}/"
        except:
            pass
    
    # Generate hashtags from mall name
    if metadata["mall_name"] and metadata["mall_name"] != "Not Available":
        mall_tag = metadata["mall_name"].replace(" ", "").replace("Mall", "")
        metadata["hashtags"] = f"#{mall_tag}Mall #Shopping #CityLife"
    
    return metadata


def _create_meta_data_tab(wb, metadata):
    """Create Mall Meta Data tab"""
    ws = wb.create_sheet("Mall Meta Data")
    
    # Header style
    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    ws["A2"] = "Meta Data"
    ws["B2"] = "Value"
    ws["A2"].fill = header_fill
    ws["B2"].fill = header_fill
    ws["A2"].font = header_font
    ws["B2"].font = header_font
    ws["A2"].alignment = center_align
    ws["B2"].alignment = center_align
    ws["A2"].border = thin_border
    ws["B2"].border = thin_border
    
    # Data rows
    rows = [
        ("Mall Name", metadata.get("mall_name", "Not Available")),
        ("Address", metadata.get("address", "Not Available")),
        ("Official Web Site", metadata.get("official_website", "Not Available")),
        ("Mall Facebook Link", metadata.get("facebook_link", "Not Available")),
        ("Mall Instagram Link", metadata.get("instagram_link", "Not Available")),
        ("Hashtags for use in Youtube , X(Twitter) Posts", metadata.get("hashtags", "#Shopping #CityLife")),
        ("Research Run Date", metadata.get("research_run_date", datetime.now().strftime("%Y-%m-%d"))),
    ]

    # Optional: scrape statistics (only shown if available)
    scrape_stats = []
    if "scraped_total_count" in metadata:
        scrape_stats.extend([
            ("Total Items Scraped (All Sources)", metadata.get("scraped_total_count", 0)),
            ("Website Tenants Scraped", metadata.get("scraped_website_count", 0)),
            ("Facebook Posts Scraped", metadata.get("scraped_facebook_count", 0)),
            ("Instagram Posts Scraped", metadata.get("scraped_instagram_count", 0)),
        ])

    rows.extend(scrape_stats)
    
    start_row = 3
    for i, (key, value) in enumerate(rows):
        row_num = start_row + i
        
        # Robustly handle list inputs (e.g. LLM returns list of hashtags)
        if isinstance(value, list):
            value = " ".join(str(v) for v in value)
        
        c1 = ws.cell(row=row_num, column=1, value=key)
        c2 = ws.cell(row=row_num, column=2, value=value)
        c1.border = thin_border
        c2.border = thin_border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60


def _create_existing_tenants_tab(wb, scraped_df, structured_data, google_search_results=None):
    """Create Existing Tennent Research tab with tenant-matched Facebook/Instagram posts.
    google_search_results: list of dicts from SERP (news/blogs) to fill column L and M for first row.
    """
    if google_search_results is None:
        google_search_results = []
    ws = wb.create_sheet("Existing Tennent Research")
    
    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers (added column M: News/Blog URL)
    headers = [
        ["Si", "Proposed Floor Number", "Proposed Shop Number", "Tennent Name", 
         "Information from Mall Website", "Facebook Scrapping", "Facebook Post URL", "Facebook Date/Time",
         "Instagram Scrapping", "Instagram Post URL", "Instagram Date/Time", "Google API Search", "News/Blog URL"]
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers[0], start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Merge headers for source columns
    ws.merge_cells('F1:F2')  # Facebook Scrapping
    ws.merge_cells('G1:G2')  # Facebook Post URL
    ws.merge_cells('H1:H2')  # Facebook Date/Time
    ws.merge_cells('I1:I2')  # Instagram Scrapping
    ws.merge_cells('J1:J2')  # Instagram Post URL
    ws.merge_cells('K1:K2')  # Instagram Date/Time
    ws.merge_cells('L1:L2')  # Google API Search
    ws.merge_cells('M1:M2')  # News/Blog URL
    
    # Write main header
    ws['B1'] = "Official Mall Directory List / Tennent Scrapping"
    ws.merge_cells('B1:C1')
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['D1'] = "Mall Directory List / Tennent Scrapping"
    ws.merge_cells('D1:E1')
    ws['D1'].fill = header_fill
    ws['D1'].font = header_font
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['F1'] = "Facebook Scrapping\nInformation from Facebook"
    ws['F1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['G1'] = "Facebook Post URL\nURL of Facebook Post"
    ws['G1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['H1'] = "Facebook Date/Time\nPost Date and Time"
    ws['H1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['I1'] = "Instagram Scrapping\nInformation from Instagram"
    ws['I1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['J1'] = "Instagram Post URL\nURL of Instagram Post"
    ws['J1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['K1'] = "Instagram Date/Time\nPost Date and Time"
    ws['K1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['L1'] = "Google API Search\nGeneral Information from Internet"
    ws['L1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws['M1'] = "News/Blog URL\nURL of News or Blog"
    ws['M1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Prepare tenant data - website data in column E, Facebook data in column F, Instagram data in column G
    tenant_data = []
    if scraped_df is not None and not scraped_df.empty:
        # Separate website, Facebook, and Instagram data
        if 'source' in scraped_df.columns:
            website_df = scraped_df[scraped_df['source'].str.lower().str.contains('website', na=False)]
            facebook_df = scraped_df[scraped_df['source'].str.lower().str.contains('facebook', na=False)]
            instagram_df = scraped_df[scraped_df['source'].str.lower().str.contains('instagram', na=False)]
        else:
            website_df = pd.DataFrame()
            facebook_df = pd.DataFrame()
            instagram_df = pd.DataFrame()
        
        # Process website data - put in website column (E)
        for idx, (_, row) in enumerate(website_df.iterrows(), start=1):
            tenant_data.append({
                'si': idx,
                'floor': row.get('floor', '-'),
                'shop_number': '-',
                'name': row.get('shop_name', ''),
                'website_info': 'Found',
                'facebook_info': '',
                'instagram_info': '',
                'instagram_datetime': '',
                'google_info': '',
                'google_url': ''
            })

        # Fallback: if no website rows, populate Existing Tenant Research from all scraped rows
        # so the sheet is never empty when there is scraped data (e.g. Facebook/Instagram only).
        if not tenant_data:
            for idx, (_, row) in enumerate(scraped_df.iterrows(), start=1):
                src = str(row.get('source', '')).lower()
                is_web = 'website' in src
                name = row.get('shop_name', '') or row.get('post_text', '') or row.get('full_text', '') or '-'
                tenant_data.append({
                    'si': idx,
                    'floor': row.get('floor', '-'),
                    'shop_number': '-',
                    'name': name,
                    'website_info': 'Found' if is_web else '',
                    'facebook_info': '',
                    'instagram_info': '',
                    'instagram_datetime': '',
                    'google_info': '',
                    'google_url': ''
                })
    
    # Assign SERP news/blog results to tenant rows (AI extraction + tenant match, else score-based)
    serp_per_tenant = []  # list of (google_info, google_url) per tenant
    if google_search_results and tenant_data:
        try:
            serp_per_tenant = _process_serp_for_tenants(google_search_results, tenant_data)
            for idx, (text, urls) in enumerate(serp_per_tenant):
                if idx < len(tenant_data) and (text or urls):
                    tenant_data[idx]["google_info"] = text
                    tenant_data[idx]["google_url"] = urls
        except Exception:
            pass
    
    # Common alignment for text cells (wrap and top-align so long text goes to next line)
    wrapped_top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Write website data first (columns A-E)
    current_row = 3
    for tenant in tenant_data:
        ws.cell(row=current_row, column=1, value=tenant['si'])  # Si
        ws.cell(row=current_row, column=2, value=tenant['floor'])  # Proposed Floor Number
        ws.cell(row=current_row, column=3, value=tenant['shop_number'])  # Proposed Shop Number

        # Tennent Name and info cells use wrapped alignment so text doesn't run horizontally forever
        name_cell = ws.cell(row=current_row, column=4, value=tenant['name'])  # Tennent Name
        name_cell.alignment = wrapped_top_align

        web_cell = ws.cell(row=current_row, column=5, value=tenant['website_info'])  # Information from Mall Website
        web_cell.alignment = wrapped_top_align

        fb_cell = ws.cell(row=current_row, column=6, value=tenant['facebook_info'])  # Facebook Scrapping (initially empty)
        fb_cell.alignment = wrapped_top_align

        ig_cell = ws.cell(row=current_row, column=9, value=tenant['instagram_info'])  # Instagram Scrapping
        ig_cell.alignment = wrapped_top_align

        ws.cell(row=current_row, column=7, value='')  # Facebook Post URL (initially empty)
        ws.cell(row=current_row, column=8, value='')  # Facebook Date/Time (initially empty)
        ws.cell(row=current_row, column=10, value='')  # Instagram Post URL (initially empty)
        ws.cell(row=current_row, column=11, value='')  # Instagram Date/Time (initially empty)
        # Google API Search (L) and News/Blog URL (M): tenant-matched SERP data (like Facebook/Instagram)
        ws.cell(row=current_row, column=12, value=tenant['google_info'])
        ws.cell(row=current_row, column=13, value=tenant.get('google_url', ''))
        ws.cell(row=current_row, column=12).alignment = wrapped_top_align
        ws.cell(row=current_row, column=13).alignment = wrapped_top_align
        current_row += 1
    
    # Match Facebook posts to tenant rows (choose BEST matching tenant per post)
    if scraped_df is not None and not scraped_df.empty:
        if 'source' in scraped_df.columns:
            facebook_df = scraped_df[scraped_df['source'].str.lower().str.contains('facebook', na=False)]

            for _, row in facebook_df.iterrows():
                # Prefer post_text, then shop_name for matching
                post_text = row.get('post_text', '') or row.get('shop_name', '')
                if not post_text:
                    continue

                best_idx = None
                best_score = 0
                for tenant_idx, tenant in enumerate(tenant_data):
                    tenant_name = tenant.get('name', '')
                    if not tenant_name:
                        continue
                    score = _score_post_for_tenant(post_text, tenant_name)
                    if score > best_score:
                        best_score = score
                        best_idx = tenant_idx

                if best_idx is not None and best_score > 0:
                    tenant_row = best_idx + 3  # +3 because first 2 rows are headers
                    existing_fb = ws.cell(row=tenant_row, column=6).value or ''
                    if existing_fb:
                        new_content = f"{existing_fb}\n\n---\n\n{post_text}"
                    else:
                        new_content = post_text
                    fb_cell = ws.cell(row=tenant_row, column=6, value=new_content)
                    fb_cell.alignment = wrapped_top_align
                    
                    # Add Facebook post URL
                    post_url = row.get('post_url', '') or ''
                    existing_fb_url = ws.cell(row=tenant_row, column=7).value or ''  # Column G: Facebook Post URL
                    if post_url:
                        if existing_fb_url:
                            new_url = f"{existing_fb_url}\n\n{post_url}"
                        else:
                            new_url = post_url
                        fb_url_cell = ws.cell(row=tenant_row, column=7, value=new_url)
                        fb_url_cell.alignment = wrapped_top_align
                    
                    # Add Facebook Date/Time (similar to Instagram)
                    post_date = row.get('post_date', '') or ''
                    if post_date:
                        # Format the date/time for display
                        date_time_display = post_date
                        try:
                            # Try to parse ISO format timestamp
                            if 'T' in post_date or post_date.endswith('Z'):
                                dt = datetime.fromisoformat(post_date.replace('Z', '+00:00'))
                                date_time_display = dt.strftime('%Y-%m-%d %H:%M:%S')
                            # If it's already a readable format, use it as-is
                        except Exception:
                            # If parsing fails, use the original value
                            pass
                        
                        existing_fb_date = ws.cell(row=tenant_row, column=8).value or ''  # Column H: Facebook Date/Time
                        if existing_fb_date:
                            new_date = f"{existing_fb_date}\n\n{date_time_display}"
                        else:
                            new_date = date_time_display
                        fb_date_cell = ws.cell(row=tenant_row, column=8, value=new_date)
                        fb_date_cell.alignment = wrapped_top_align
                
    # Match Instagram posts to tenant rows (choose BEST matching tenant per post)
    if scraped_df is not None and not scraped_df.empty:
        if 'source' in scraped_df.columns:
            instagram_df = scraped_df[scraped_df['source'].str.lower().str.contains('instagram', na=False)]

            for _, row in instagram_df.iterrows():
                # Prefer full_text, then shop_name
                post_text = row.get('full_text', '') or row.get('shop_name', '')
                if not post_text:
                    continue

                time_text = row.get('time', '')
                datetime_val = row.get('datetime', '')

                # Format date/time
                date_time_display = ''
                if datetime_val:
                    try:
                        dt = datetime.fromisoformat(datetime_val.replace('Z', '+00:00'))
                        date_time_display = dt.strftime('%Y-%m-%d %H:%M:%S')
                        if time_text:
                            date_time_display += f' ({time_text})'
                    except Exception:
                        if time_text and datetime_val:
                            date_time_display = f"{time_text} | {datetime_val}"
                        elif datetime_val:
                            date_time_display = datetime_val
                        elif time_text:
                            date_time_display = time_text
                elif time_text:
                    date_time_display = time_text

                best_idx = None
                best_score = 0
                for tenant_idx, tenant in enumerate(tenant_data):
                    tenant_name = tenant.get('name', '')
                    if not tenant_name:
                        continue
                    score = _score_post_for_tenant(post_text, tenant_name)
                    if score > best_score:
                        best_score = score
                        best_idx = tenant_idx

                if best_idx is not None and best_score > 0:
                    tenant_row = best_idx + 3
                    existing_ig = ws.cell(row=tenant_row, column=9).value or ''  # Column I: Instagram Scrapping
                    if existing_ig:
                        new_content = f"{existing_ig}\n\n---\n\n{post_text}"
                    else:
                        new_content = post_text
                    ig_cell = ws.cell(row=tenant_row, column=9, value=new_content)  # Column I: Instagram Scrapping
                    ig_cell.alignment = wrapped_top_align

                    # Add Instagram post URL
                    post_url = row.get('post_url', '') or ''
                    existing_ig_url = ws.cell(row=tenant_row, column=10).value or ''  # Column J: Instagram Post URL
                    if post_url:
                        if existing_ig_url:
                            new_url = f"{existing_ig_url}\n\n{post_url}"
                        else:
                            new_url = post_url
                        ig_url_cell = ws.cell(row=tenant_row, column=10, value=new_url)
                        ig_url_cell.alignment = wrapped_top_align

                    existing_dt = ws.cell(row=tenant_row, column=11).value or ''  # Column K: Instagram Date/Time
                    if existing_dt and date_time_display:
                        ws.cell(row=tenant_row, column=11, value=f"{existing_dt}\n{date_time_display}")
                    elif date_time_display:
                        ws.cell(row=tenant_row, column=11, value=date_time_display)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30  # Facebook Scrapping
    ws.column_dimensions['G'].width = 50  # Facebook Post URL
    ws.column_dimensions['H'].width = 30  # Facebook Date/Time
    ws.column_dimensions['I'].width = 30  # Instagram Scrapping
    ws.column_dimensions['J'].width = 50  # Instagram Post URL
    ws.column_dimensions['K'].width = 25  # Instagram Date/Time
    ws.column_dimensions['L'].width = 40  # Google API Search
    ws.column_dimensions['M'].width = 50  # News/Blog URL


def _create_coming_soon_tab(wb, structured_data, coming_soon_shops=None):
    """Create Coming Soon Tennent Research tab with real data extracted from website.
    
    Args:
        wb: Workbook object
        structured_data: Comparison data structure (not used for coming soon)
        coming_soon_shops: List of shop names that are coming soon (extracted from website)
    """
    ws = wb.create_sheet("Coming Soon Tennent Research")
    
    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Simple header - just "Coming Soon" column
    ws.merge_cells('A1:B1')
    cell = ws.cell(row=1, column=1, value="Coming Soon Shops")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Column headers
    ws.cell(row=2, column=1, value="Si")
    ws.cell(row=2, column=2, value="Coming Soon")
    
    # Style header row
    for col in [1, 2]:
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Use real coming soon shops if provided, otherwise show message
    if coming_soon_shops and len(coming_soon_shops) > 0:
        # Write real coming soon shops data
        for idx, shop_name in enumerate(coming_soon_shops, start=1):
            row = idx + 2
            ws.cell(row=row, column=1, value=idx)  # Si
            ws.cell(row=row, column=2, value=shop_name)  # Coming Soon
            
            # Add borders
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
    else:
        # No coming soon shops found - show message
        row = 3
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="No coming soon shops found on the website.")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50


def _is_likely_tenant_name(name):
    """Return True if name looks like a real shop/tenant name, False if it looks like FB/IG post content."""
    if not name or not isinstance(name, str):
        return False
    s = name.strip()
    if len(s) < 2:
        return False
    # Exclude obvious social media content
    s_lower = s.lower()
    if any(x in s_lower for x in (
        'instagram.com', 'facebook.com', 'fb.com', 'instagr.am',
        'reel', '| ', ' |', 'sponsored by @', 'http://', 'https://',
        'video sponsored', 'others |', ' | 3w |', ' | 1w |',
        'post ', 'caption', '.reel', ' insta', ' at westfield', ' at mall',
        ' others ', 'buy ', 'get 1 free', 'get one free', '... and '
    )):
        return False
    # Exclude very long strings (post captions, not tenant names)
    if len(s) > 50:
        return False
    return True


def _create_vacated_shops_tab(wb, structured_data):
    """Create Vacated Shops tab showing shops that were in old data but missing in new data.
    Uses ONLY website/directory tenant data for comparison — Facebook and Instagram are excluded."""
    ws = wb.create_sheet("Vacated Shops")
    
    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Header - vacated shops based on website tenant list only
    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1, value="Vacated Shops (Website directory only — shops in old data but missing from current website tenant list. Facebook/Instagram not used.)")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Column headers
    headers = ["Si", "Shop Name", "Phone", "Floor"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Extract vacated shops from structured_data (website-only comparison)
    # vacated_shops at top level come from compare_shops(old_df, website_df, website_only=True)
    vacated_shops = []
    if structured_data and isinstance(structured_data, dict):
        raw = structured_data.get("vacated_shops", [])
        # First filter: exclude obvious Facebook/Instagram content (post captions, URLs, etc.)
        vacated_shops = [s for s in raw if _is_likely_tenant_name(s.get("shop_name", ""))]
    
    # Validate shop names using AI to filter out non-shop entries (Facebook/Instagram post text, etc.)
    validated_vacated_shops = []
    if vacated_shops:
        try:
            from llm_engine import validate_shop_names
            # Extract shop names for validation
            shop_names = [shop.get("shop_name", "") for shop in vacated_shops if shop.get("shop_name")]
            
            if shop_names:
                print(f"Validating {len(shop_names)} vacated shop names using AI...")
                validated_names = validate_shop_names(shop_names)
                print(f"AI validated {len(validated_names)} real shop names out of {len(shop_names)} entries")
                
                # Create a set of validated names for quick lookup
                validated_set = {name.lower().strip() for name in validated_names}
                
                # Keep only shops with validated names
                for shop in vacated_shops:
                    shop_name = shop.get("shop_name", "").strip()
                    if shop_name and shop_name.lower() in validated_set:
                        validated_vacated_shops.append(shop)
        except Exception as e:
            print(f"Warning: Failed to validate shop names with AI: {e}, using all shops")
            # Fallback: use all shops if validation fails
            validated_vacated_shops = vacated_shops
    
    # Use validated shops
    vacated_shops = validated_vacated_shops
    
    # If no vacated shops, show a message
    if not vacated_shops:
        row = 3
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws.cell(row=row, column=1, value="No vacated shops found. All shops from old data are still present in the website directory.")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    else:
        # Write validated vacated shops data
        for idx, shop in enumerate(vacated_shops, start=1):
            row = idx + 2
            ws.cell(row=row, column=1, value=idx)  # Si (serial number)
            ws.cell(row=row, column=2, value=shop.get("shop_name", ""))  # Shop Name
            ws.cell(row=row, column=3, value=shop.get("phone", ""))  # Phone
            ws.cell(row=row, column=4, value=shop.get("floor", ""))  # Floor
            
            # Add borders to all cells
            for col in [1, 2, 3, 4]:
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def _create_ai_analysis_tab(wb, llm_json, structured_data=None):
    """Create AI Analysis Report tab"""
    ws = wb.create_sheet("AI Analysis Report")
    
    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    row = 1
    
    if not llm_json:
        ws.cell(row=row, column=1, value="No AI analysis available")
        return
    
    # Facebook Data Report
    if "facebook" in llm_json:
        fb_data = llm_json["facebook"]
        
        # Header
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="FACEBOOK DATA REPORT")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Data rows
        ws.cell(row=row, column=1, value="Occupancy Trend")
        ws.cell(row=row, column=2, value=str(fb_data.get("occupancy_trend", "")))
        row += 1
        
        ws.cell(row=row, column=1, value="New Shops")
        new_shops = fb_data.get("new_shops", "")
        ws.cell(row=row, column=2, value=str(new_shops) if new_shops else "")
        row += 1
        
        ws.cell(row=row, column=1, value="Vacancy Changes")
        vacancy = fb_data.get("vacancy_changes", "")
        ws.cell(row=row, column=2, value=str(vacancy) if isinstance(vacancy, bool) else str(vacancy))
        row += 1
        
        ws.cell(row=row, column=1, value="Business Insights")
        insights = fb_data.get("business_insights", [])
        if insights:
            insights_text = "• " + "\n• ".join(insights) if isinstance(insights, list) else str(insights)
        else:
            insights_text = ""
        ws.cell(row=row, column=2, value=insights_text)
        row += 1
        
        row += 1  # Blank row
    
    # Website Data Report
    if "website" in llm_json:
        web_data = llm_json["website"]
        
        # Header
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="WEBSITE DATA REPORT")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Data rows
        ws.cell(row=row, column=1, value="Occupancy Trend")
        ws.cell(row=row, column=2, value=str(web_data.get("occupancy_trend", "")))
        row += 1
        
        ws.cell(row=row, column=1, value="New Shops")
        new_shops = web_data.get("new_shops", "")
        ws.cell(row=row, column=2, value=str(new_shops) if new_shops else "")
        row += 1
        
        ws.cell(row=row, column=1, value="Vacancy Changes")
        vacancy = web_data.get("vacancy_changes", "")
        ws.cell(row=row, column=2, value=str(vacancy) if isinstance(vacancy, bool) else str(vacancy))
        row += 1
        
        ws.cell(row=row, column=1, value="Business Insights")
        insights = web_data.get("business_insights", [])
        if insights:
            insights_text = "• " + "\n• ".join(insights) if isinstance(insights, list) else str(insights)
        else:
            insights_text = ""
        ws.cell(row=row, column=2, value=insights_text)
        row += 1
        
        row += 1  # Blank row
    
    # Instagram Data Report
    if "instagram" in llm_json:
        ig_data = llm_json["instagram"]
        
        # Header
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="INSTAGRAM DATA REPORT")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Data rows
        ws.cell(row=row, column=1, value="Occupancy Trend")
        ws.cell(row=row, column=2, value=str(ig_data.get("occupancy_trend", "")))
        row += 1
        
        ws.cell(row=row, column=1, value="New Shops")
        new_shops = ig_data.get("new_shops", "")
        ws.cell(row=row, column=2, value=str(new_shops) if new_shops else "")
        row += 1
        
        ws.cell(row=row, column=1, value="Vacancy Changes")
        vacancy = ig_data.get("vacancy_changes", "")
        ws.cell(row=row, column=2, value=str(vacancy) if isinstance(vacancy, bool) else str(vacancy))
        row += 1
        
        ws.cell(row=row, column=1, value="Business Insights")
        insights = ig_data.get("business_insights", [])
        if insights:
            insights_text = "• " + "\n• ".join(insights) if isinstance(insights, list) else str(insights)
        else:
            insights_text = ""
        ws.cell(row=row, column=2, value=insights_text)
        row += 1
        
        row += 1  # Blank row
    
    # Add New Shops List Section if structured_data is available
    if structured_data and 'new_shops' in structured_data:
        new_shops = structured_data.get('new_shops', [])
        if new_shops:
            row += 2  # Add blank rows
            
            # Header for New Shops List
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws.cell(row=row, column=1, value="NEW SHOPS LIST")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
            
            # List new shops (tenant name only)
            for shop in new_shops:
                shop_name = shop.get('shop_name', '') or ''
                ws.cell(row=row, column=1, value="•")
                ws.cell(row=row, column=2, value=shop_name)
                row += 1
                row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80


def _parse_post_date_for_sort(date_str):
    """Parse post date string to datetime for sorting. Returns naive datetime (datetime.min for missing/invalid)."""
    if not date_str or (isinstance(date_str, float) and pd.isna(date_str)) or str(date_str).strip() in ('', 'nan', '-'):
        return datetime.min
    try:
        s = str(date_str).strip()
        if 'T' in s or s.endswith('Z'):
            dt = datetime.fromisoformat(s.replace('Z', '+00:00'))
        elif len(s) >= 10:
            dt = datetime.fromisoformat(s[:19] if len(s) > 19 else s)
        else:
            return datetime.min
        return dt.replace(tzinfo=None) if dt.tzinfo else dt
    except Exception:
        return datetime.min


def _create_facebook_scratch_tab(wb, scraped_df):
    """Create Facebook Scratch tab with all Facebook posts (SN, Date, Post, URL). Sorted by date, latest first."""
    ws = wb.create_sheet("Facebook Scratch")
    
    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["SN", "Date", "Post", "Post URL"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Extract Facebook posts from scraped_df
    facebook_posts = []
    if scraped_df is not None and not scraped_df.empty:
        if 'source' in scraped_df.columns:
            facebook_df = scraped_df[scraped_df['source'].str.lower().str.contains('facebook', na=False)]
            
            for idx, (_, row) in enumerate(facebook_df.iterrows(), start=1):
                # Get post text - prefer 'post_text' column, fallback to 'shop_name'
                post_text = ''
                if 'post_text' in row and pd.notna(row.get('post_text')):
                    post_text = str(row.get('post_text', ''))
                if not post_text and 'shop_name' in row:
                    post_text = str(row.get('shop_name', ''))
                
                # Get post date - prefer 'post_date' column, fallback to empty
                post_date = ''
                if 'post_date' in row and pd.notna(row.get('post_date')):
                    post_date = str(row.get('post_date', ''))
                
                # Format date if it's an ISO timestamp
                if post_date and post_date.strip() and post_date != 'nan':
                    try:
                        # Try to parse ISO format timestamp
                        if 'T' in post_date or post_date.endswith('Z'):
                            dt = datetime.fromisoformat(post_date.replace('Z', '+00:00'))
                            post_date = dt.strftime('%Y-%m-%d %H:%M:%S')
                        elif len(post_date) > 10:
                            # Try other date formats
                            dt = datetime.fromisoformat(post_date)
                            post_date = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        # If parsing fails, use the date as-is (might be a readable string)
                        pass
                else:
                    post_date = '-'

                # Get post URL - prefer explicit 'post_url' column, fallback to phone if it looks like a URL
                post_url = ''
                if 'post_url' in row and pd.notna(row.get('post_url')):
                    post_url = str(row.get('post_url', ''))
                elif 'phone' in row and pd.notna(row.get('phone')):
                    possible_url = str(row.get('phone', ''))
                    if 'http' in possible_url or 'facebook.com' in possible_url or possible_url.startswith('www.'):
                        post_url = possible_url

                raw_date = row.get('post_date', '')
                facebook_posts.append({
                    'sn': idx,
                    'date': post_date if post_date else '-',
                    'post': post_text if post_text else '-',
                    'url': post_url if post_url else '-',
                    '_sort_key': _parse_post_date_for_sort(raw_date if pd.notna(raw_date) else '')
                })
    
    # Sort by date, latest first
    if facebook_posts:
        facebook_posts.sort(key=lambda p: p['_sort_key'], reverse=True)
    
    # Write data rows (already sorted latest first)
    if facebook_posts:
        for row_idx, post in enumerate(facebook_posts, start=2):
            # SN (re-number 1, 2, 3... after sort)
            cell_sn = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            cell_sn.border = thin_border
            cell_sn.alignment = center_align
            
            # Date
            cell_date = ws.cell(row=row_idx, column=2, value=post['date'])
            cell_date.border = thin_border
            cell_date.alignment = Alignment(horizontal="left", vertical="top")
            
            # Post
            cell_post = ws.cell(row=row_idx, column=3, value=post['post'])
            cell_post.border = thin_border
            cell_post.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Post URL
            cell_url = ws.cell(row=row_idx, column=4, value=post['url'])
            cell_url.border = thin_border
            cell_url.alignment = Alignment(horizontal="left", vertical="top")
    else:
        # No Facebook posts found
        ws.cell(row=2, column=1, value="No Facebook posts found")
        ws.merge_cells('A2:D2')
        cell = ws.cell(row=2, column=1)
        cell.alignment = center_align
        cell.border = thin_border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 10  # SN
    ws.column_dimensions['B'].width = 20  # Date
    ws.column_dimensions['C'].width = 80  # Post
    ws.column_dimensions['D'].width = 60  # Post URL


def _create_instagram_scratch_tab(wb, scraped_df):
    """Create Instagram Scratch tab with all Instagram posts (SN, Date/Time, Post, URL). Sorted by date, latest first."""
    ws = wb.create_sheet("Instagram Scratch")

    # Header style (dark maroon background)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["SN", "Date/Time", "Post", "Post URL"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    instagram_posts = []
    if scraped_df is not None and not scraped_df.empty:
        if 'source' in scraped_df.columns:
            instagram_df = scraped_df[scraped_df['source'].str.lower().str.contains('instagram', na=False)]

            for idx, (_, row) in enumerate(instagram_df.iterrows(), start=1):
                # Post text - prefer full_text, then shop_name
                post_text = ''
                if 'full_text' in row and pd.notna(row.get('full_text')):
                    post_text = str(row.get('full_text', ''))
                if not post_text and 'shop_name' in row:
                    post_text = str(row.get('shop_name', ''))

                # Date/time - prefer datetime, then time, then post_date
                post_dt = ''
                raw_date_for_sort = ''
                if 'datetime' in row and pd.notna(row.get('datetime')):
                    post_dt = str(row.get('datetime', ''))
                    raw_date_for_sort = post_dt
                elif 'time' in row and pd.notna(row.get('time')):
                    post_dt = str(row.get('time', ''))
                    raw_date_for_sort = post_dt
                elif 'post_date' in row and pd.notna(row.get('post_date')):
                    post_dt = str(row.get('post_date', ''))
                    raw_date_for_sort = post_dt

                # Format datetime if ISO
                if post_dt and post_dt.strip() and post_dt != 'nan':
                    try:
                        if 'T' in post_dt or post_dt.endswith('Z'):
                            dt = datetime.fromisoformat(post_dt.replace('Z', '+00:00'))
                            post_dt = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        pass
                else:
                    post_dt = '-'

                # Post URL - prefer post_url, then phone if it's a URL
                post_url = ''
                if 'post_url' in row and pd.notna(row.get('post_url')):
                    post_url = str(row.get('post_url', ''))
                elif 'phone' in row and pd.notna(row.get('phone')):
                    possible_url = str(row.get('phone', ''))
                    if 'http' in possible_url or 'instagram.com' in possible_url or possible_url.startswith('www.'):
                        post_url = possible_url

                instagram_posts.append({
                    'sn': idx,
                    'date': post_dt if post_dt else '-',
                    'post': post_text if post_text else '-',
                    'url': post_url if post_url else '-',
                    '_sort_key': _parse_post_date_for_sort(raw_date_for_sort)
                })

    # Sort by date, latest first
    if instagram_posts:
        instagram_posts.sort(key=lambda p: p['_sort_key'], reverse=True)

    # Write data rows (already sorted latest first)
    if instagram_posts:
        for row_idx, post in enumerate(instagram_posts, start=2):
            # SN (re-number 1, 2, 3... after sort)
            cell_sn = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            cell_sn.border = thin_border
            cell_sn.alignment = center_align

            # Date/Time
            cell_date = ws.cell(row=row_idx, column=2, value=post['date'])
            cell_date.border = thin_border
            cell_date.alignment = Alignment(horizontal="left", vertical="top")

            # Post
            cell_post = ws.cell(row=row_idx, column=3, value=post['post'])
            cell_post.border = thin_border
            cell_post.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Post URL
            cell_url = ws.cell(row=row_idx, column=4, value=post['url'])
            cell_url.border = thin_border
            cell_url.alignment = Alignment(horizontal="left", vertical="top")
    else:
        ws.cell(row=2, column=1, value="No Instagram posts found")
        ws.merge_cells('A2:D2')
        cell = ws.cell(row=2, column=1)
        cell.alignment = center_align
        cell.border = thin_border

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 10  # SN
    ws.column_dimensions['B'].width = 25  # Date/Time
    ws.column_dimensions['C'].width = 80  # Post
    ws.column_dimensions['D'].width = 60  # Post URL


def _create_serp_scratch_tab(wb, google_search_results):
    """Create Google SERP Scratch tab with all SERP API results (same structure as terminal output).
    Columns: SN, Title, General Information (snippet), URL, Source, Date.
    """
    ws = wb.create_sheet("Google SERP Scratch")
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["SN", "Title", "General Information", "URL", "Source", "Date"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    if google_search_results:
        for row_idx, item in enumerate(google_search_results, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=(item.get("title") or "").strip())
            ws.cell(row=row_idx, column=3, value=(item.get("snippet") or "").strip())
            ws.cell(row=row_idx, column=4, value=(item.get("link") or "").strip())
            ws.cell(row=row_idx, column=5, value=(item.get("source") or "").strip())
            ws.cell(row=row_idx, column=6, value=(item.get("date") or "").strip())
            for c in range(1, 7):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    else:
        ws.cell(row=2, column=1, value="No SERP API data found")
        ws.merge_cells('A2:F2')
        cell = ws.cell(row=2, column=1)
        cell.alignment = center_align
        cell.border = thin_border

    ws.column_dimensions['A'].width = 8   # SN
    ws.column_dimensions['B'].width = 35  # Title
    ws.column_dimensions['C'].width = 60  # General Information
    ws.column_dimensions['D'].width = 55  # URL
    ws.column_dimensions['E'].width = 20  # Source
    ws.column_dimensions['F'].width = 18  # Date
