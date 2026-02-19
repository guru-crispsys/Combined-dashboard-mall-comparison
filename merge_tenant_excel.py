"""
Merge Mall AI Dashboard tenant CSV with Map Scraping Excel report.

- Matches rows by tenant name (Excel "Tennent Name" <-> CSV "name").
- Replaces "Proposed Floor Number" in Excel with CSV "floor".
- Fills "Proposed Shop Number" with CSV "location_id".
- Inserts "Latitude" and "Longitude" columns after "Proposed Shop Number".
- Adds a new tab "Mall Data Match Status" with all mall CSV rows and a Match Status column (Matched/Mismatched).
"""

from __future__ import annotations

import io
import re
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# Sheet and column names we expect in the Excel report
EXISTING_TENANTS_SHEET = "Existing Tennent Research"
MALL_DATA_MATCH_SHEET = "Mall Data Match Status"
HEADER_ROW = 2  # 1-based; row 2 has Si, Proposed Floor Number, etc.
DATA_START_ROW = 3  # 1-based

# Column letters / positions in the ORIGINAL Excel (before we insert Lat/Long)
COL_SI = 1
COL_PROPOSED_FLOOR = 2
COL_PROPOSED_SHOP = 3
COL_TENANT_NAME = 4
# After inserting 2 columns at 4: Lat=4, Long=5, Tenant Name moves to 6, etc.
COL_LATITUDE_NEW = 4
COL_LONGITUDE_NEW = 5
COL_TENANT_NAME_AFTER_INSERT = 6


def _normalize_name(name: Optional[str]) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).strip().lower()
    # Collapse multiple spaces and remove extra punctuation for matching
    s = re.sub(r"\s+", " ", s)
    return s


def _build_name_lookup(csv_df: pd.DataFrame) -> dict[str, dict]:
    """Build a lookup: normalized_name -> {floor, location_id, latitude, longitude}."""
    lookup = {}
    required = ["name", "floor", "location_id", "latitude", "longitude"]
    for c in required:
        if c not in csv_df.columns:
            raise ValueError(f"CSV must contain column: {c}")
    for _, row in csv_df.iterrows():
        name = row.get("name")
        n = _normalize_name(name)
        if not n:
            continue
        # Prefer first occurrence; optionally you could prefer non-null lat/long
        if n not in lookup:
            lookup[n] = {
                "floor": row.get("floor"),
                "location_id": row.get("location_id"),
                "latitude": row.get("latitude"),
                "longitude": row.get("longitude"),
            }
    return lookup


def merge_tenant_csv_with_excel(
    csv_content: bytes | io.BytesIO,
    excel_content: bytes | io.BytesIO,
) -> bytes:
    """
    Merge mall tenants CSV into the Excel report.

    - Reads CSV (name, floor, location_id, latitude, longitude).
    - Opens Excel and finds sheet "Existing Tennent Research".
    - Inserts Latitude and Longitude columns after Proposed Shop Number.
    - For each data row, matches by tenant name and fills:
      Proposed Floor Number <- floor,
      Proposed Shop Number <- location_id,
      Latitude, Longitude <- latitude, longitude.

    Returns the merged Excel file as bytes.
    """
    if isinstance(csv_content, bytes):
        csv_content = io.BytesIO(csv_content)
    if isinstance(excel_content, bytes):
        excel_content = io.BytesIO(excel_content)

    csv_df = pd.read_csv(csv_content, encoding="utf-8")
    name_lookup = _build_name_lookup(csv_df)

    wb = load_workbook(excel_content, read_only=False, data_only=False)
    if EXISTING_TENANTS_SHEET not in wb.sheetnames:
        raise ValueError(f"Excel must contain a sheet named '{EXISTING_TENANTS_SHEET}'")
    ws = wb[EXISTING_TENANTS_SHEET]

    # Insert two columns after "Proposed Shop Number" (column C = index 3 → insert at 4)
    ws.insert_cols(COL_LATITUDE_NEW, 2)

    # Set headers for the new columns (row 2)
    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(row=HEADER_ROW, column=COL_LATITUDE_NEW, value="Latitude")
    ws.cell(row=HEADER_ROW, column=COL_LATITUDE_NEW).fill = header_fill
    ws.cell(row=HEADER_ROW, column=COL_LATITUDE_NEW).font = header_font
    ws.cell(row=HEADER_ROW, column=COL_LATITUDE_NEW).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=HEADER_ROW, column=COL_LONGITUDE_NEW, value="Longitude")
    ws.cell(row=HEADER_ROW, column=COL_LONGITUDE_NEW).fill = header_fill
    ws.cell(row=HEADER_ROW, column=COL_LONGITUDE_NEW).font = header_font
    ws.cell(row=HEADER_ROW, column=COL_LONGITUDE_NEW).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge B1:C1 stays; we need to extend merge for "Official Mall Directory..." to include new columns if desired.
    # Optional: merge D1:E1 for "Latitude / Longitude" - leaving as separate columns per requirement.

    # Set of normalized tenant names that exist in the Excel report (for Match Status tab)
    excel_tenant_names_normalized = set()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        tenant_cell = ws.cell(row=row, column=COL_TENANT_NAME_AFTER_INSERT)
        tenant_name_raw = tenant_cell.value
        norm = _normalize_name(tenant_name_raw)
        if norm:
            excel_tenant_names_normalized.add(norm)
        info = name_lookup.get(norm) if norm else None
        if not info:
            continue
        # Replace Proposed Floor Number (col 2)
        ws.cell(row=row, column=COL_PROPOSED_FLOOR, value=info["floor"])
        # Fill Proposed Shop Number (col 3) with location_id
        loc_id = info["location_id"]
        if pd.notna(loc_id):
            ws.cell(row=row, column=COL_PROPOSED_SHOP, value=loc_id)
        # New columns: Latitude (4), Longitude (5)
        lat, lon = info["latitude"], info["longitude"]
        if pd.notna(lat):
            ws.cell(row=row, column=COL_LATITUDE_NEW, value=lat)
        if pd.notna(lon):
            ws.cell(row=row, column=COL_LONGITUDE_NEW, value=lon)

    # New tab: all mall data from CSV + Match Status (Matched / Mismatched)
    csv_content.seek(0)
    csv_df_full = pd.read_csv(csv_content, encoding="utf-8")
    match_status = [
        "Matched" if _normalize_name(name) in excel_tenant_names_normalized else "Mismatched"
        for name in csv_df_full.get("name", pd.Series(dtype=object))
    ]
    csv_df_full["Match Status"] = match_status

    # Create new sheet (place after Existing Tennent Research)
    if MALL_DATA_MATCH_SHEET in wb.sheetnames:
        del wb[MALL_DATA_MATCH_SHEET]
    idx = wb.sheetnames.index(EXISTING_TENANTS_SHEET) + 1
    ws_mall = wb.create_sheet(MALL_DATA_MATCH_SHEET, idx)

    header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(csv_df_full.columns, start=1):
        cell = ws_mall.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(csv_df_full.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_mall.cell(row=r_idx, column=c_idx, value=val)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
