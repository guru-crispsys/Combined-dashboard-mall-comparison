"""
Full retail store-opening discovery pipeline.

Flow: Query generation → Selenium search → Requests (with Selenium fallback) → BeautifulSoup
      → Gemini AI analysis → Structured output → CSV/Excel export + extracted text files.
"""

import csv
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import EXTRACTED_OUTPUT_DIR, STRUCTURED_OUTPUT_DIR

from ai_analysis import AI_AVAILABLE, AI_SOURCE_NAME, analyze_extracted_text, extract_combined, generate_mall_intel
from extract_text import extract_clean_text, extract_text_from_url
from query_generation import extract_mall_name_from_query, generate_queries
from selenium_search import create_driver, extract_ai_overview, find_official_mall_website, search_google


def _sanitize_filename(s: str, max_len: int = 60) -> str:
    import re
    s = re.sub(r"[^\w\s\-.]", "", s)
    s = re.sub(r"[\s_\-]+", "_", s).strip("_")
    return (s[:max_len] if len(s) > max_len else s) or "page"


def run_pipeline_gemini_only(
    mall_name: Optional[str] = None,
    brand_name: Optional[str] = None,
    custom_query: Optional[str] = None,
    export_csv: bool = True,
    export_excel: bool = True,
    save_extracted_text: bool = True,
) -> Dict[str, Any]:
    """
    Gemini-only pipeline: send user prompt to Gemini, parse response into structured output.
    No Selenium or web scraping.

    Steps:
    1. Build user prompt from mall_name / brand_name / custom_query.
    2. Call Gemini to generate mall/retail intel (one generate_content call).
    3. Parse Gemini's text with extract_combined → store_openings, vacated_tenants, temporary_events, latest_updates.
    4. Dedupe and optionally export CSV/Excel.

    Returns:
        Dict with "store_openings", "vacated_tenants", "temporary_events", "latest_updates".
    """
    queries = generate_queries(mall_name=mall_name, brand_name=brand_name, custom_query=custom_query)
    user_prompt = queries[0] if queries else "Coming soon shops and events at malls"
    print(f"[AI-only] User prompt: {user_prompt}")

    raw_text = generate_mall_intel(user_prompt, debug=True)
    if not raw_text or not raw_text.strip():
        print("[AI-only] No response from AI.")
        return {"store_openings": [], "vacated_tenants": [], "temporary_events": [], "latest_updates": [], "extracted_text_files": []}

    slug = _sanitize_filename(user_prompt[:50])
    extracted_content = f"Prompt: {user_prompt}\n" + "=" * 70 + "\n\n" + raw_text
    extracted_filename = f"extract_ai_{slug}.txt"
    extracted_text_files = [{"filename": extracted_filename, "content": extracted_content}]

    if save_extracted_text:
        out_dir = Path(EXTRACTED_OUTPUT_DIR)
        out_dir.mkdir(exist_ok=True)
        fpath = out_dir / extracted_filename
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(extracted_content)
        print(f"[AI-only] Saved response to {fpath.name}")

    result = extract_combined(raw_text, source_url="", source_title=AI_SOURCE_NAME, debug=True)
    structured_rows = result.get("store_openings") or []
    vacated_tenants_list = result.get("vacated_tenants") or []
    temporary_events_list = result.get("temporary_events") or []
    latest_update_one = result.get("latest_updates")

    # Dedupe store_openings by (mall_name, brand_name)
    seen_key: set = set()
    unique_rows = []
    for row in structured_rows:
        key = (row.get("mall_name", ""), row.get("brand_name", ""))
        if key in seen_key:
            continue
        seen_key.add(key)
        unique_rows.append(row)

    # Dedupe vacated_tenants by (mall_name, brand_name)
    seen_vacated: set = set()
    unique_vacated = []
    for v in vacated_tenants_list:
        key = (v.get("mall_name", ""), v.get("brand_name", ""))
        if key in seen_vacated:
            continue
        seen_vacated.add(key)
        unique_vacated.append(v)

    # Dedupe temporary_events by (mall_name, event_name, date_or_range)
    seen_events: set = set()
    unique_events = []
    for e in temporary_events_list:
        key = (e.get("mall_name", ""), e.get("event_name", ""), e.get("date_or_range", ""))
        if key in seen_events:
            continue
        seen_events.add(key)
        unique_events.append(e)

    unique_updates = [latest_update_one] if latest_update_one else []

    # Export CSV/Excel to folder only when requested
    if export_csv or export_excel:
        structured_dir = Path(STRUCTURED_OUTPUT_DIR)
        structured_dir.mkdir(exist_ok=True)
    if export_csv and unique_rows:
        csv_path = structured_dir / "store_openings.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            fieldnames = ["mall_name", "brand_name", "expected_opening", "location_context", "confidence", "source_url", "source_title"]
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(unique_rows)
        print(f"Exported CSV: {csv_path}")

    if export_excel and unique_rows:
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Store openings"
            headers = ["Mall", "Brand", "Expected Opening", "Location Context", "Confidence", "Source URL", "Source Title"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, row in enumerate(unique_rows, 2):
                ws.cell(row=row_idx, column=1, value=row.get("mall_name", ""))
                ws.cell(row=row_idx, column=2, value=row.get("brand_name", ""))
                ws.cell(row=row_idx, column=3, value=row.get("expected_opening", ""))
                ws.cell(row=row_idx, column=4, value=row.get("location_context", ""))
                ws.cell(row=row_idx, column=5, value=row.get("confidence", ""))
                ws.cell(row=row_idx, column=6, value=row.get("source_url", ""))
                ws.cell(row=row_idx, column=7, value=row.get("source_title", ""))
            xlsx_path = structured_dir / "store_openings.xlsx"
            wb.save(xlsx_path)
            print(f"Exported Excel: {xlsx_path}")
        except ImportError:
            pass

    return {"store_openings": unique_rows, "vacated_tenants": unique_vacated, "temporary_events": unique_events, "latest_updates": unique_updates, "extracted_text_files": extracted_text_files}


def run_pipeline(
    mall_name: Optional[str] = None,
    brand_name: Optional[str] = None,
    custom_query: Optional[str] = None,
    max_links_per_query: int = 5,
    max_results_per_search: int = 15,
    skip_ai_relevance_check: bool = False,
    export_csv: bool = True,
    export_excel: bool = True,
    save_extracted_text: bool = True,
) -> Dict[str, Any]:
    """
    Run the full pipeline and return structured store-opening records.

    Steps:
    1. Generate search queries from mall_name / brand_name / custom_query.
    2. For each query: Selenium search.
    3. For each result link: fetch page, extract text with BeautifulSoup.
    4. Optionally save extracted text to files.
    5. Run Gemini: relevance check + structured extraction (Mall, Brand, Expected Opening, Location, Confidence).
    6. Export to CSV and Excel.

    Returns:
        Dict with "store_openings" (list of mall/brand/opening records) and "latest_updates" (list of mall update records).
    """
    queries = generate_queries(mall_name=mall_name, brand_name=brand_name, custom_query=custom_query)
    print(f"[Step 1] Generated {len(queries)} query(s): {queries}")

    structured_rows: List[Dict[str, Any]] = []
    vacated_tenants_list: List[Dict[str, Any]] = []
    temporary_events_list: List[Dict[str, Any]] = []
    latest_updates_list: List[Dict[str, Any]] = []
    out_dir = Path(EXTRACTED_OUTPUT_DIR)
    structured_dir = Path(STRUCTURED_OUTPUT_DIR)
    driver = create_driver()
    try:
        all_results: List[dict] = []  # search result items {title, link, snippet}
        seen_links: set = set()

        # --- Functionality 1: If query is about a mall, find and scrape official mall website first ---
        if custom_query and custom_query.strip():
            mall_from_query = extract_mall_name_from_query(custom_query)
            if mall_from_query:
                print(f"[Step 1b] Detected mall: «{mall_from_query}» — finding official website...")
                official = find_official_mall_website(mall_from_query, driver, max_results=10)
                if official:
                    off_url = official.get("link") or ""
                    off_title = official.get("title") or off_url
                    print(f"[Step 1b] Official site: {off_title[:60]}...")
                    text = extract_text_from_url(off_url)
                    if not text:
                        try:
                            driver.get(off_url)
                            time.sleep(1)
                            text = extract_clean_text(driver.page_source or "")
                        except Exception as e:
                            print(f"[Step 1b] Could not load official site: {e}")
                    if text:
                        source_label = f"Official website: {mall_from_query}"
                        result = analyze_extracted_text(
                            text,
                            source_url=off_url,
                            source_title=source_label,
                            skip_relevance_check=True,
                            debug=True,
                        )
                        for row in result.get("store_openings") or []:
                            structured_rows.append(row)
                        for row in result.get("vacated_tenants") or []:
                            vacated_tenants_list.append(row)
                        for row in result.get("temporary_events") or []:
                            temporary_events_list.append(row)
                        if result.get("latest_updates"):
                            latest_updates_list.append(result["latest_updates"])
                        if save_extracted_text:
                            out_dir.mkdir(exist_ok=True)
                            slug = _sanitize_filename(mall_from_query)
                            fpath = out_dir / f"extract_official_{slug}.txt"
                            with open(fpath, "w", encoding="utf-8") as f:
                                f.write(f"URL: {off_url}\nTitle: {source_label}\n")
                                f.write("=" * 70 + "\n\n")
                                f.write(text)
                            print(f"[Step 1b] Saved official site text: {fpath.name}")
                        seen_links.add(off_url)
                        print(f"[Step 1b] Extracted coming soon / latest updates from official site first.")
                    else:
                        print(f"[Step 1b] No text extracted from official site (skipped).")
                else:
                    print(f"[Step 1b] No official mall website found for «{mall_from_query}».")

        print("[Step 2] Running Selenium search (including AI Overview / Dive deeper)...")
        ai_overview_sources: List[Dict[str, Any]] = []  # [{query, text, source_url, source_title}]
        for q in queries:
            results = search_google(q, max_results=max_results_per_search, driver=driver)
            print(f"  Query '{q[:50]}...' -> {len(results)} result(s)")
            # Extract Google AI Overview / AI mode text from current search page (waits for async load)
            try:
                ai_data = extract_ai_overview(
                    driver,
                    expand_first=True,
                    wait_after_load=2.5,
                    initial_wait=5.0,  # AI Overview loads asynchronously; wait for it
                )
                ai_text = (ai_data.get("text") or "").strip()
                if ai_text:
                    ai_overview_sources.append({
                        "query": q,
                        "text": ai_text,
                        "source_url": driver.current_url,
                        "source_title": f"Google AI Overview: {q[:60]}",
                    })
                    print(f"  AI Overview extracted: {len(ai_text)} chars")
            except Exception as e:
                print(f"  AI Overview skip: {e}")
            for r in results:
                link = r.get("link") or ""
                if link and link not in seen_links:
                    seen_links.add(link)
                    all_results.append(r)
            if len(all_results) >= max_links_per_query * 3:  # cap total links
                break

        print(f"[Step 2] Total unique links to process: {len(all_results)} | AI Overviews: {len(ai_overview_sources)}")
        if not all_results:
            print("[FAIL] No search results. Check Selenium/Chrome or try a different query.")
            return {"store_openings": [], "vacated_tenants": [], "temporary_events": [], "latest_updates": [], "extracted_text_files": []}

        # Limit how many pages we fetch
        to_process = all_results[: max_links_per_query * len(queries)]
        to_process = to_process[:20]

        if save_extracted_text:
            out_dir.mkdir(exist_ok=True)
        extracted_text_files: List[Dict[str, str]] = []
        file_index = 0

        # Process Google AI Overview text first (Dive deeper / AI mode)
        for ai_src in ai_overview_sources:
            file_index += 1
            text = ai_src.get("text") or ""
            link = ai_src.get("source_url") or ""
            title = ai_src.get("source_title") or "Google AI Overview"
            if not text:
                continue
            print(f"\n  [AI Overview {file_index}] {title[:60]}...")
            print(f"      Source: Google AI Overview ({len(text)} chars)")
            slug = _sanitize_filename(ai_src.get("query", "ai_overview"))
            filename = f"extract_ai_{file_index}_{slug}.txt"
            content = f"URL: {link}\nTitle: {title}\n" + "=" * 70 + "\n\n" + text
            extracted_text_files.append({"filename": filename, "content": content})
            if save_extracted_text:
                fpath = out_dir / filename
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(content)
                print(f"      Saved text: {fpath.name}")
            result = analyze_extracted_text(
                text,
                source_url=link,
                source_title=title,
                skip_relevance_check=True,
                debug=True,
            )
            for row in result.get("store_openings") or []:
                structured_rows.append(row)
            for row in result.get("vacated_tenants") or []:
                vacated_tenants_list.append(row)
            for row in result.get("temporary_events") or []:
                temporary_events_list.append(row)
            if result.get("latest_updates"):
                latest_updates_list.append(result["latest_updates"])

        print("[Step 3–5] Fetching pages, extracting text, running AI analysis...")
        for i, r in enumerate(to_process):
            link = r.get("link") or ""
            title = r.get("title") or link
            file_index += 1
            print(f"\n  [{file_index}/{len(to_process)}] {title[:60]}...")
            print(f"      URL: {link[:70]}...")
            text = extract_text_from_url(link)
            if not text:
                print(f"      Requests empty -> trying Selenium fallback...")
                try:
                    driver.get(link)
                    time.sleep(2)
                    text = extract_clean_text(driver.page_source or "")
                except Exception as e:
                    print(f"      Selenium fallback failed: {e}")
                if not text:
                    print(f"      SKIP: Could not fetch or extract text (empty).")
                    continue
                print(f"      (Selenium fallback) Text length: {len(text)} chars")
            else:
                print(f"      Text length: {len(text)} chars")
        if save_extracted_text:
            slug = _sanitize_filename(title or link)
            fpath = out_dir / f"extract_{file_index}_{slug}.txt"
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(f"URL: {link}\nTitle: {title}\n")
                f.write("=" * 70 + "\n\n")
                f.write(text)
            print(f"      Saved text: {fpath.name}")

        result = analyze_extracted_text(
            text,
            source_url=link,
            source_title=title,
            skip_relevance_check=skip_ai_relevance_check,
            debug=True,
        )
        rows = result.get("store_openings") or []
        vacated = result.get("vacated_tenants") or []
        update = result.get("latest_updates")
        if rows:
            print(f"      -> {len(rows)} store-opening row(s) extracted")
        if vacated:
            print(f"      -> {len(vacated)} vacated tenant(s) extracted")
        for row in rows:
            structured_rows.append(row)
        for row in vacated:
            vacated_tenants_list.append(row)
        for row in result.get("temporary_events") or []:
            temporary_events_list.append(row)
        if update:
            latest_updates_list.append(update)

    finally:
        driver.quit()

    print(f"\n[Step 6] Total structured rows before dedupe: {len(structured_rows)}")
    # Dedupe by (mall_name, brand_name) keeping first
    seen_key: set = set()
    unique_rows = []
    for row in structured_rows:
        key = (row.get("mall_name", ""), row.get("brand_name", ""))
        if key in seen_key:
            continue
        seen_key.add(key)
        unique_rows.append(row)
    print(f"[Step 6] After dedupe: {len(unique_rows)} unique row(s).")

    # Dedupe latest_updates by source_url
    seen_urls: set = set()
    unique_updates = []
    for u in latest_updates_list:
        url = u.get("source_url") or ""
        if url in seen_urls:
            continue
        seen_urls.add(url)
        unique_updates.append(u)
    print(f"[Step 6] Latest updates: {len(unique_updates)} unique update(s).")

    # Dedupe vacated_tenants by (mall_name, brand_name)
    seen_vacated: set = set()
    unique_vacated = []
    for v in vacated_tenants_list:
        key = (v.get("mall_name", ""), v.get("brand_name", ""))
        if key in seen_vacated:
            continue
        seen_vacated.add(key)
        unique_vacated.append(v)
    print(f"[Step 6] Vacated tenants: {len(unique_vacated)} unique.")

    # Dedupe temporary_events by (mall_name, event_name, date_or_range)
    seen_events: set = set()
    unique_events = []
    for e in temporary_events_list:
        key = (e.get("mall_name", ""), e.get("event_name", ""), e.get("date_or_range", ""))
        if key in seen_events:
            continue
        seen_events.add(key)
        unique_events.append(e)
    print(f"[Step 6] Temporary events: {len(unique_events)} unique.")

    # Export CSV
    if export_csv and unique_rows:
        csv_path = structured_dir / "store_openings.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            fieldnames = ["mall_name", "brand_name", "expected_opening", "location_context", "confidence", "source_url", "source_title"]
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(unique_rows)
        print(f"Exported CSV: {csv_path}")

    # Export Excel
    if export_excel and unique_rows:
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Store openings"
            headers = ["Mall", "Brand", "Expected Opening", "Location Context", "Confidence", "Source URL", "Source Title"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, row in enumerate(unique_rows, 2):
                ws.cell(row=row_idx, column=1, value=row.get("mall_name", ""))
                ws.cell(row=row_idx, column=2, value=row.get("brand_name", ""))
                ws.cell(row=row_idx, column=3, value=row.get("expected_opening", ""))
                ws.cell(row=row_idx, column=4, value=row.get("location_context", ""))
                ws.cell(row=row_idx, column=5, value=row.get("confidence", ""))
                ws.cell(row=row_idx, column=6, value=row.get("source_url", ""))
                ws.cell(row=row_idx, column=7, value=row.get("source_title", ""))
            xlsx_path = structured_dir / "store_openings.xlsx"
            wb.save(xlsx_path)
            print(f"Exported Excel: {xlsx_path}")
        except ImportError:
            print("Install openpyxl for Excel export: pip install openpyxl")

    return {"store_openings": unique_rows, "vacated_tenants": unique_vacated, "temporary_events": unique_events, "latest_updates": unique_updates}


if __name__ == "__main__":
    import sys

    # CMD: python pipeline.py "custom query"   OR   python pipeline.py MallName BrandName
    # Uses Gemini only (no Selenium).
    custom = None
    mall = None
    brand = None
    if len(sys.argv) >= 2:
        arg1 = sys.argv[1].strip()
        if len(sys.argv) >= 3:
            mall = arg1
            brand = sys.argv[2].strip()
        else:
            custom = arg1

    print("Running Gemini-only pipeline (no web search)...")
    if not AI_AVAILABLE:
        print("Warning: No AI API available. Set OPENAI_API_KEY or GEMINI_API_KEY.")
    out = run_pipeline_gemini_only(
        mall_name=mall,
        brand_name=brand,
        custom_query=custom,
        save_extracted_text=True,
        export_csv=True,
        export_excel=True,
    )
    rows = out.get("store_openings") or []
    vacated = out.get("vacated_tenants") or []
    events = out.get("temporary_events") or []
    updates = out.get("latest_updates") or []
    print(f"\nNew tenants / store openings: {len(rows)}")
    for r in rows:
        print(f"  - {r.get('mall_name')} | {r.get('brand_name')} | {r.get('expected_opening')} | {r.get('confidence')}")
    print(f"\nVacated tenants: {len(vacated)}")
    for v in vacated:
        print(f"  - {v.get('mall_name')} | {v.get('brand_name')} | {v.get('closed_date')} | {v.get('notes', '')[:50]}")
    print(f"\nTemporary events: {len(events)}")
    for e in events:
        print(f"  - {e.get('mall_name')} | {e.get('event_name')} | {e.get('date_or_range')} | {e.get('event_type')}")
    print(f"\nLatest updates: {len(updates)}")
    for u in updates:
        print(f"  - {u.get('mall_name')}: {u.get('hours_weather', '')[:60]}...")
